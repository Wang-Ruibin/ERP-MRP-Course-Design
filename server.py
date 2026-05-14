import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

import mysql.connector
from flask import Flask, jsonify, request, send_file, send_from_directory
from mysql.connector import Error
from openpyxl import Workbook
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "db_config.json"
SQL_SCHEMA_PATH = BASE_DIR / "mysql_schema.sql"

app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path="")


def load_db_config():
    with CONFIG_PATH.open("r", encoding="utf-8") as config_file:
        return json.load(config_file)


def get_connection(include_database=True):
    config = load_db_config().copy()
    database = config.pop("database", None)
    if include_database and database:
        config["database"] = database
    return mysql.connector.connect(**config)


def execute_sql_script():
    script = SQL_SCHEMA_PATH.read_text(encoding="utf-8")
    statements = []
    current = []

    for line in script.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("--"):
            continue
        current.append(line)
        if stripped.endswith(";"):
            statements.append("\n".join(current))
            current = []

    if current:
        statements.append("\n".join(current))

    connection = get_connection(include_database=False)
    try:
        cursor = connection.cursor()
        for statement in statements:
            cursor.execute(statement)
        connection.commit()
    finally:
        cursor.close()
        connection.close()


def fetch_state():
    connection = get_connection(include_database=True)
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT item_code, item_name, lead_time_weeks, on_hand_qty, allocated_qty,
                   safety_stock, lot_rule, lot_size
            FROM items
            ORDER BY item_name, item_code
            """
        )
        items = [
            {
                "code": row["item_code"],
                "name": row["item_name"],
                "leadTime": int(row["lead_time_weeks"]),
                "onHand": int(row["on_hand_qty"]),
                "allocated": int(row["allocated_qty"]),
                "safetyStock": int(row["safety_stock"]),
                "lotRule": row["lot_rule"],
                "lotSize": int(row["lot_size"]),
            }
            for row in cursor.fetchall()
        ]

        cursor.execute(
            """
            SELECT p.item_code AS parent_code, c.item_code AS child_code, b.quantity_per_parent
            FROM bom_relations b
            JOIN items p ON p.id = b.parent_item_id
            JOIN items c ON c.id = b.child_item_id
            ORDER BY p.item_name, c.item_name, b.id
            """
        )
        bom = [
            {
                "parent": row["parent_code"],
                "child": row["child_code"],
                "quantity": int(row["quantity_per_parent"]),
            }
            for row in cursor.fetchall()
        ]

        cursor.execute(
            """
            SELECT i.item_code, d.demand_week, d.demand_qty
            FROM demand_orders d
            JOIN items i ON i.id = d.item_id
            ORDER BY i.item_name, d.demand_week, d.id
            """
        )
        demands = [
            {
                "itemCode": row["item_code"],
                "week": int(row["demand_week"]),
                "quantity": int(row["demand_qty"]),
            }
            for row in cursor.fetchall()
        ]

        horizon = max(
            [12]
            + [demand["week"] for demand in demands]
            + [item["leadTime"] + 8 for item in items]
        )

        return {
            "horizon": min(52, horizon),
            "items": items,
            "bom": bom,
            "demands": demands,
        }
    finally:
        cursor.close()
        connection.close()


def save_state(payload):
    data = payload.get("data", {})
    items = normalize_items(data.get("items", []))
    valid_codes = {item["code"] for item in items}
    bom = normalize_bom(data.get("bom", []), valid_codes)
    demands = normalize_demands(data.get("demands", []), valid_codes)

    connection = get_connection(include_database=True)
    cursor = connection.cursor()
    try:
        connection.start_transaction()
        cursor.execute("DELETE FROM demand_orders")
        cursor.execute("DELETE FROM scheduled_receipts")
        cursor.execute("DELETE FROM bom_relations")
        cursor.execute("DELETE FROM items")

        item_id_map = {}
        for item in items:
            cursor.execute(
                """
                INSERT INTO items
                (item_code, item_name, lead_time_weeks, on_hand_qty, allocated_qty, safety_stock, lot_rule, lot_size)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    item["code"],
                    item["name"],
                    item["leadTime"],
                    item["onHand"],
                    item["allocated"],
                    item["safetyStock"],
                    item["lotRule"],
                    item["lotSize"],
                ),
            )
            item_id_map[item["code"]] = cursor.lastrowid

        for relation in bom:
            cursor.execute(
                """
                INSERT INTO bom_relations (parent_item_id, child_item_id, quantity_per_parent)
                VALUES (%s, %s, %s)
                """,
                (
                    item_id_map[relation["parent"]],
                    item_id_map[relation["child"]],
                    relation["quantity"],
                ),
            )

        for demand in demands:
            cursor.execute(
                """
                INSERT INTO demand_orders (item_id, demand_week, demand_qty)
                VALUES (%s, %s, %s)
                """,
                (
                    item_id_map[demand["itemCode"]],
                    demand["week"],
                    demand["quantity"],
                ),
            )

        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()


def normalize_items(raw_items):
    cleaned = []
    seen = set()
    for item in raw_items:
        code = str(item.get("code", "")).strip().upper()
        if not code or code in seen:
            continue
        seen.add(code)
        cleaned.append(
            {
                "code": code,
                "name": str(item.get("name", "")).strip() or code,
                "leadTime": clamp_int(item.get("leadTime"), 0),
                "onHand": clamp_int(item.get("onHand"), 0),
                "allocated": clamp_int(item.get("allocated"), 0),
                "safetyStock": clamp_int(item.get("safetyStock"), 0),
                "lotRule": "FIXED" if str(item.get("lotRule")) == "FIXED" else "L4L",
                "lotSize": clamp_int(item.get("lotSize"), 0),
            }
        )
    return cleaned


def normalize_bom(raw_bom, valid_codes):
    cleaned = []
    seen = set()
    for relation in raw_bom:
        parent = str(relation.get("parent", "")).strip().upper()
        child = str(relation.get("child", "")).strip().upper()
        if not parent or not child or parent == child:
            continue
        if parent not in valid_codes or child not in valid_codes:
            continue
        key = (parent, child)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(
            {
                "parent": parent,
                "child": child,
                "quantity": max(1, clamp_int(relation.get("quantity"), 1)),
            }
        )
    return cleaned


def normalize_demands(raw_demands, valid_codes):
    cleaned = []
    for demand in raw_demands:
        item_code = str(demand.get("itemCode", "")).strip().upper()
        if item_code not in valid_codes:
            continue
        quantity = clamp_int(demand.get("quantity"), 0)
        if quantity <= 0:
            continue
        cleaned.append(
            {
                "itemCode": item_code,
                "week": max(1, clamp_int(demand.get("week"), 1)),
                "quantity": quantity,
            }
        )
    return cleaned


def clamp_int(value, fallback):
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return fallback


def save_history_snapshot(data, result):
    payload = json.dumps({"data": data, "result": result}, ensure_ascii=False)
    report_name = f"MRP_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    item_count = len(data.get("items", []))
    order_count = len(result.get("orders", []))

    connection = get_connection(include_database=True)
    cursor = connection.cursor()
    try:
        cursor.execute(
            """
            INSERT INTO report_history (report_name, item_count, order_count, snapshot_json)
            VALUES (%s, %s, %s, %s)
            """,
            (report_name, item_count, order_count, payload),
        )
        connection.commit()
        return report_name
    finally:
        cursor.close()
        connection.close()


def list_history():
    connection = get_connection(include_database=True)
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT id, report_name, item_count, order_count,
                   DATE_FORMAT(created_at, '%Y-%m-%d %H:%i:%s') AS created_at
            FROM report_history
            ORDER BY id DESC
            LIMIT 100
            """
        )
        return [
            {
                "id": row["id"],
                "reportName": row["report_name"],
                "itemCount": int(row["item_count"]),
                "orderCount": int(row["order_count"]),
                "createdAt": row["created_at"],
            }
            for row in cursor.fetchall()
        ]
    finally:
        cursor.close()
        connection.close()


def get_history_snapshot(history_id):
    connection = get_connection(include_database=True)
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT id, report_name, snapshot_json
            FROM report_history
            WHERE id = %s
            """,
            (history_id,),
        )
        row = cursor.fetchone()
        if not row:
            raise ValueError("未找到对应的历史记录。")
        snapshot = json.loads(row["snapshot_json"])
        return row["report_name"], snapshot
    finally:
        cursor.close()
        connection.close()


def delete_history_snapshot(history_id):
    connection = get_connection(include_database=True)
    cursor = connection.cursor()
    try:
        cursor.execute("DELETE FROM report_history WHERE id = %s", (history_id,))
        connection.commit()
        if cursor.rowcount == 0:
            raise ValueError("未找到对应的历史记录。")
    finally:
        cursor.close()
        connection.close()


def build_workbook(data, result):
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "订货建议"
    detail_sheet = workbook.create_sheet("MRP明细")

    bold_font = Font(bold=True)
    summary_rows = [
        ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["计划周期", result.get("horizon", data.get("horizon", 12))],
        ["物料数", len(data.get("items", []))],
        ["订货建议数", len(result.get("orders", []))],
    ]
    for row in summary_rows:
        orders_sheet.append(row)
    orders_sheet.append([])

    order_headers = ["物料编码", "物料名称", "计划下达周", "计划接收周", "订货量", "来源"]
    orders_sheet.append(order_headers)
    for cell in orders_sheet[6]:
        cell.font = bold_font

    for order in result.get("orders", []):
        orders_sheet.append(
            [
                order.get("itemCode", ""),
                order.get("itemName", ""),
                format_week(order.get("releaseWeek", 0)),
                format_week(order.get("receiptWeek", 0)),
                order.get("quantity", 0),
                order.get("source", ""),
            ]
        )

    horizon = int(result.get("horizon", data.get("horizon", 12)))
    week_headers = ["项目", "提前"] + [f"W{index}" for index in range(1, horizon + 1)]

    for item in result.get("items", []):
        code = item.get("code", "")
        matrix = (result.get("resultByItem") or {}).get(code, {})
        detail_sheet.append([f"{code} / {item.get('name', '')}"])
        detail_sheet[detail_sheet.max_row][0].font = bold_font
        detail_sheet.append(
            [
                f"LT={item.get('leadTime', 0)} 周",
                f"批量={item.get('lotRule', 'L4L')}",
                f"固定批量={item.get('lotSize', 0)}",
            ]
        )
        detail_sheet.append(week_headers)
        for cell in detail_sheet[detail_sheet.max_row]:
            cell.font = bold_font

        append_detail_row(detail_sheet, "毛需求", matrix.get("gross", []), horizon)
        append_detail_row(detail_sheet, "预计可得", matrix.get("projected", []), horizon)
        append_detail_row(detail_sheet, "净需求", matrix.get("net", []), horizon)
        append_detail_row(detail_sheet, "计划接收", matrix.get("receipts", []), horizon)
        append_detail_row(detail_sheet, "计划下达", matrix.get("releases", []), horizon)
        detail_sheet.append([])

    autosize_sheet(orders_sheet)
    autosize_sheet(detail_sheet)
    return workbook


def append_detail_row(sheet, label, values, horizon):
    normalized = list(values or [])
    while len(normalized) <= horizon:
        normalized.append(0)
    row = [label]
    for index in range(0, horizon + 1):
        value = normalized[index]
        row.append(value if value else "")
    sheet.append(row)


def autosize_sheet(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 36)


def format_week(week):
    return "第 1 周前" if int(week) == 0 else f"第 {int(week)} 周"


@app.get("/api/health")
def health():
    config = load_db_config()
    try:
        connection = get_connection(include_database=True)
        connection.close()
        return jsonify(
            {
                "ok": True,
                "database_connected": True,
                "database": config["database"],
                "message": "MySQL connection ok",
            }
        )
    except Error as error:
        try:
            fallback = get_connection(include_database=False)
            fallback.close()
            return (
                jsonify(
                    {
                        "ok": False,
                        "database_connected": False,
                        "database": config["database"],
                        "message": f"数据库服务可连接，但库或表未就绪：{error}",
                    }
                ),
                500,
            )
        except Error:
            return (
                jsonify(
                    {
                        "ok": False,
                        "database_connected": False,
                        "database": config["database"],
                        "message": str(error),
                    }
                ),
                500,
            )


@app.post("/api/init-db")
def init_db():
    try:
        execute_sql_script()
        return jsonify({"ok": True, "data": fetch_state()})
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.get("/api/data")
def get_data():
    try:
        return jsonify({"ok": True, "data": fetch_state()})
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.post("/api/data")
def post_data():
    payload = request.get_json(silent=True) or {}
    try:
        save_state(payload)
        return jsonify({"ok": True})
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.get("/api/history")
def get_history():
    try:
        return jsonify({"ok": True, "history": list_history()})
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.get("/api/history/<int:history_id>")
def get_history_detail(history_id):
    try:
        report_name, snapshot = get_history_snapshot(history_id)
        return jsonify(
            {
                "ok": True,
                "reportName": report_name,
                "data": snapshot.get("data", {}),
                "result": snapshot.get("result", {}),
            }
        )
    except ValueError as error:
        return jsonify({"ok": False, "message": str(error)}), 404
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.delete("/api/history/<int:history_id>")
def delete_history_detail(history_id):
    try:
        delete_history_snapshot(history_id)
        return jsonify({"ok": True})
    except ValueError as error:
        return jsonify({"ok": False, "message": str(error)}), 404
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.post("/api/export-excel")
def export_excel():
    payload = request.get_json(silent=True) or {}
    data = payload.get("data", {})
    result = payload.get("result", {})
    if not data or not result:
        return jsonify({"ok": False, "message": "导出数据不完整。"}), 400

    try:
        report_name = save_history_snapshot(data, result)
        workbook = build_workbook(data, result)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename = f"{report_name}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Error as error:
        return jsonify({"ok": False, "message": str(error)}), 500


@app.route("/")
def root():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/<path:path>")
def static_files(path):
    return send_from_directory(BASE_DIR, path)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8000, debug=False)
